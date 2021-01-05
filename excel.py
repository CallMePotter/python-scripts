import openpyxl

def getDuplicatesWithCount(listOfElems):
    dictOfElems = dict()
    for elem in listOfElems:
        if elem in dictOfElems:
            dictOfElems[elem] += 1
        else:
            dictOfElems[elem] = 1

    dictOfElems = { key:value for key, value in dictOfElems.items() if value > 1}
    return dictOfElems

# Define range
row_start = 1
row_end = 100

col_start = 1
col_end = 100

# Open document
doc_name = "meh2.xlsx"
document = openpyxl.load_workbook(doc_name)
sheet = document["Лист1"]

value = ""

for x in range(row_start, row_end):
    for y in range(col_start, col_end):
        cell = sheet.cell(row = x, column = y).value
        if cell != None and isinstance(cell, int) and cell > 800000:
            cell_val = sheet.cell(row = x - 1, column = y).value
            value += str(cell) + "=" + str(cell_val) + ";"

vals = value.split(";")
vals.pop()
vals.sort()

dictOfElems = getDuplicatesWithCount(vals)

print(doc_name, "\n")
for key, value in dictOfElems.items():
    print(key[:-2], str(value),int(key[-1]) * value)

print(sheet["e47"].value)
